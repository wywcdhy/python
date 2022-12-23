#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/11/4 16:50
# @Author  : wy
# @Site    : 
# @File    : excel_test.py
import openpyxl


def csv_handle():
    filename = "all.xlsx"
    wb = openpyxl.load_workbook(filename)
    sheets = wb.sheetnames    #拿到表名
    sheet_prize = wb[sheets[0]]  #拿到组内的表名
    sheet_visitor = wb[sheets[1]]
    for r in range(1, sheet_prize.max_row + 1):
        prize_value = str(sheet_prize.cell(row=r, column=2).value)
        # print(str(sheet_prize.cell(row=r, column=2).value))
        for i in range(1, sheet_visitor.max_row + 1):
            vistor_value = sheet_visitor.cell(row=i, column=1).value
            if vistor_value == prize_value:
                print(str(vistor_value) + "匹配成功！")
                mobile = str(sheet_visitor.cell(row=i, column=19).value)   #可将某值调出查看，row'行'，column'列'
                company = str(sheet_visitor.cell(row=i, column=6).value)
                jobTitle = str(sheet_visitor.cell(row=i, column=7).value)
                line_company = 'I' + str(r)
                line_mobile = 'J' + str(r)
                line_jobTitle = 'K' + str(r)
                sheet_prize[line_mobile] = mobile       #直接赋值给表中的J列r行
                sheet_prize[line_company] = company
                sheet_prize[line_jobTitle] = jobTitle
                print(sheet_prize[line_mobile].value)
                break
            else:
                print(str(prize_value) + '' + "没有该用户\n")
                continue
    print('完成匹配，正在保存！')
    wb.save(filename=filename)   #记住一定要保存


if __name__ == '__main__':
    csv_handle()
